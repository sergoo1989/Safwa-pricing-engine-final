from __future__ import annotations

from dataclasses import asdict
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from .advanced_pricing import calculate_price_breakdown
from .costing import build_cost_details, resolve_component_cost


REGULAR_PRICE_COL = "سعر المنتج"
DISCOUNT_PRICE_COL = "السعر المخفض"
FINAL_EXCL_VAT_COL = "Final_Price_Excl_VAT"
FINAL_INCL_VAT_COL = "السعر_الظاهر_للعميل_شامل_الضريبة"
PRICE_SOURCE_COL = "Price_Source"
SALLA_OUTPUT_FILE_NAME = "salla_prices_with_customer_final_price.xlsx"
SALLA_MATCH_SKU_COL = "Salla_SKU_For_Matching"
SALLA_ID_CANDIDATES = ("No.", "No", "رقم المنتج", "معرف المنتج")
SALLA_NAME_CANDIDATES = ("أسم المنتج", "اسم المنتج", "اسم المنتج في سلة")
SALLA_SKU_CANDIDATES = (
    SALLA_MATCH_SKU_COL,
    "رمز المنتج sku",
    "رمز المنتج SKU",
    "SKU",
    "sku",
    "Sku",
    "رمز المنتج",
    "رمز المنتج (SKU)",
    "رمز التخزين",
    "رمز التخزين SKU",
    "كود المنتج",
    "Product SKU",
    "Product_SKU",
)
CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp1256", "latin1")


def normalize_header(value: object) -> str:
    if pd.isna(value):
        return ""
    return " ".join(str(value).strip().split())


def unwrap_excel_text_formula(value: object) -> object:
    if not isinstance(value, str):
        return value

    text = value.strip()
    if len(text) >= 3 and text.startswith('="') and text.endswith('"'):
        return text[2:-1].replace('""', '"')
    if len(text) >= 3 and text.startswith("='") and text.endswith("'"):
        return text[2:-1]
    return value


def normalize_sku(value: object) -> str:
    if pd.isna(value):
        return ""

    value = unwrap_excel_text_formula(value)

    if isinstance(value, float) and value.is_integer():
        value = int(value)

    sku = str(value).strip()
    if sku.endswith(".0"):
        sku = sku[:-2]
    if sku.lower() in {"", "0", "0.0", "nan", "none", "null", "-"}:
        return ""
    return sku.upper()


def _source_to_bytes(source: Any) -> bytes | None:
    if isinstance(source, (str, Path)):
        return None
    if hasattr(source, "getvalue"):
        return source.getvalue()
    if hasattr(source, "seek"):
        source.seek(0)
    return source.read()


def _rewind(source: Any) -> None:
    if hasattr(source, "seek"):
        source.seek(0)


def read_excel_preserve_text_formulas(source: Any, header: int | None = 0, nrows: int | None = None) -> pd.DataFrame:
    source_bytes = _source_to_bytes(source)
    workbook_source = BytesIO(source_bytes) if source_bytes is not None else source
    _rewind(workbook_source)

    workbook = load_workbook(workbook_source, read_only=True, data_only=False)
    worksheet = workbook.worksheets[0]
    rows = []

    for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
        if nrows is not None and row_index >= nrows:
            break
        rows.append([unwrap_excel_text_formula(cell) for cell in row])

    workbook.close()

    if header is None:
        return pd.DataFrame(rows)

    if header >= len(rows):
        return pd.DataFrame()

    columns = rows[header]
    data = rows[header + 1 :]
    return pd.DataFrame(data, columns=columns)


def read_csv_with_fallback(source: Any, **kwargs) -> tuple[pd.DataFrame, str]:
    last_error: Exception | None = None
    source_bytes = _source_to_bytes(source)

    for encoding in CSV_ENCODINGS:
        try:
            if source_bytes is None:
                return pd.read_csv(source, encoding=encoding, dtype=object, **kwargs), encoding
            return pd.read_csv(BytesIO(source_bytes), encoding=encoding, dtype=object, **kwargs), encoding
        except UnicodeDecodeError as exc:
            last_error = exc

    raise ValueError(f"Could not read CSV file with supported encodings: {CSV_ENCODINGS}") from last_error


def _source_suffix(source: Any) -> str:
    name = getattr(source, "name", source)
    return Path(str(name)).suffix.lower()


def find_header_row(source: Any) -> tuple[int, str | None]:
    suffix = _source_suffix(source)

    if suffix in {".xlsx", ".xlsm"}:
        _rewind(source)
        preview = read_excel_preserve_text_formulas(source, header=None, nrows=30)
        encoding = None
    elif suffix == ".xls":
        _rewind(source)
        preview = pd.read_excel(source, header=None, nrows=30, dtype=object)
        encoding = None
    elif suffix == ".csv":
        preview, encoding = read_csv_with_fallback(source, header=None, nrows=30)
    else:
        raise ValueError("Unsupported file type. Please provide an Excel file (.xlsx/.xls) or a CSV file.")

    required = {REGULAR_PRICE_COL, DISCOUNT_PRICE_COL}
    for row_index, row in preview.iterrows():
        headers = {normalize_header(cell) for cell in row.tolist()}
        if required.issubset(headers):
            return int(row_index), encoding

    raise ValueError(
        "Required columns were not found in the first 30 rows: "
        f"{REGULAR_PRICE_COL!r}, {DISCOUNT_PRICE_COL!r}"
    )


def load_salla_export(source: Any) -> pd.DataFrame:
    header_row, encoding = find_header_row(source)
    suffix = _source_suffix(source)

    if suffix in {".xlsx", ".xlsm"}:
        _rewind(source)
        return read_excel_preserve_text_formulas(source, header=header_row)

    if suffix == ".xls":
        _rewind(source)
        return pd.read_excel(source, header=header_row, dtype=object)

    if suffix == ".csv":
        if encoding is None:
            raise ValueError("CSV encoding could not be detected.")
        source_bytes = _source_to_bytes(source)
        if source_bytes is None:
            return pd.read_csv(source, header=header_row, encoding=encoding, dtype=object)
        return pd.read_csv(BytesIO(source_bytes), header=header_row, encoding=encoding, dtype=object)

    raise ValueError("Unsupported file type. Please provide an Excel file (.xlsx/.xls) or a CSV file.")


def find_column(df: pd.DataFrame, candidates: tuple[str, ...], required: bool = True) -> str | None:
    column_lookup = {normalize_header(column): column for column in df.columns}
    for candidate in candidates:
        if candidate in column_lookup:
            return column_lookup[candidate]
    if required:
        available_columns = ", ".join(str(column) for column in df.columns)
        raise ValueError(
            "Missing required column. Expected one of: "
            + ", ".join(repr(column) for column in candidates)
            + f"\nAvailable columns: {available_columns}"
        )
    return None


def count_valid_sku_values(df: pd.DataFrame, sku_col: str | None) -> int:
    if not sku_col or sku_col not in df.columns:
        return 0
    return int(df[sku_col].map(normalize_sku).astype(bool).sum())


def detect_salla_sku_column(df: pd.DataFrame) -> str | None:
    best_col = None
    best_count = -1
    for candidate in SALLA_SKU_CANDIDATES:
        col = find_column(df, (candidate,), required=False)
        if not col:
            continue
        valid_count = count_valid_sku_values(df, col)
        if valid_count > best_count:
            best_col = col
            best_count = valid_count
    return best_col


def set_salla_match_sku_column(df: pd.DataFrame, sku_col: str) -> pd.DataFrame:
    if sku_col not in df.columns:
        raise ValueError(f"Selected SKU column does not exist: {sku_col}")
    result = df.copy()
    result[SALLA_MATCH_SKU_COL] = result[sku_col].map(normalize_sku)
    return result


def require_price_columns(df: pd.DataFrame) -> tuple[str, str]:
    regular_col = find_column(df, (REGULAR_PRICE_COL,))
    discount_col = find_column(df, (DISCOUNT_PRICE_COL,))
    return regular_col or REGULAR_PRICE_COL, discount_col or DISCOUNT_PRICE_COL


def clean_price_series(series: pd.Series) -> pd.Series:
    translation = str.maketrans(
        "٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹",
        "01234567890123456789",
    )

    def clean_one(value: object) -> object:
        if pd.isna(value):
            return pd.NA

        text = str(value).strip().translate(translation)
        if not text:
            return pd.NA

        text = text.replace("٫", ".").replace("٬", ",").replace(" ", "")
        text = "".join(char for char in text if char.isdigit() or char in ".,-")

        if not text or text in {"-", ".", ","}:
            return pd.NA

        if "," in text and "." in text:
            if text.rfind(",") > text.rfind("."):
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")
        elif "," in text:
            parts = text.split(",")
            last_part = parts[-1]
            leading = "".join(parts[:-1])
            if len(parts) > 2 and len(last_part) in {1, 2}:
                text = leading + "." + last_part
            elif len(last_part) == 3 and leading:
                text = leading + last_part
            else:
                text = leading + "." + last_part if leading else last_part

        return text

    return pd.to_numeric(series.map(clean_one), errors="coerce")


def add_customer_final_price_columns(df: pd.DataFrame, vat_rate: float = 0.15) -> tuple[pd.DataFrame, dict[str, int]]:
    regular_col, discount_col = require_price_columns(df)

    result = df.copy()
    regular_price = clean_price_series(result[regular_col])
    discount_price = clean_price_series(result[discount_col])

    uses_discount = discount_price.notna() & (discount_price > 0)
    final_excl_vat = regular_price.where(~uses_discount, discount_price)
    final_incl_vat = (final_excl_vat * (1 + vat_rate)).round(2)

    result[FINAL_EXCL_VAT_COL] = final_excl_vat
    result[FINAL_INCL_VAT_COL] = final_incl_vat
    result[PRICE_SOURCE_COL] = "Regular Price"
    result.loc[uses_discount, PRICE_SOURCE_COL] = "Discount Price"

    summary = {
        "total_rows": int(len(result)),
        "discount_rows": int(uses_discount.sum()),
        "regular_rows": int((~uses_discount).sum()),
        "missing_final_price_rows": int(final_excl_vat.isna().sum()),
    }
    return result, summary


def get_saved_salla_path(data_dir: str | Path) -> Path:
    return Path(data_dir) / SALLA_OUTPUT_FILE_NAME


def save_salla_prices(data_dir: str | Path, df: pd.DataFrame) -> Path:
    output_path = get_saved_salla_path(data_dir)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False)
    return output_path


def load_saved_salla_prices(data_dir: str | Path) -> pd.DataFrame:
    path = get_saved_salla_path(data_dir)
    if not path.exists():
        raise FileNotFoundError(f"Salla prices file is not saved yet: {path}")
    return pd.read_excel(path, dtype=object)


def _build_cost_name_lookup(
    materials: dict,
    products_summary: pd.DataFrame,
    packages_summary: pd.DataFrame,
) -> dict[str, str]:
    names: dict[str, str] = {}

    for sku, material in (materials or {}).items():
        sku = str(sku).strip()
        if not sku:
            continue
        names[sku] = str(getattr(material, "material_name", "") or sku).strip()

    if products_summary is not None and not products_summary.empty:
        for _, row in products_summary.iterrows():
            sku = str(row.get("Product_SKU", "") or "").strip()
            if sku:
                names[sku] = str(row.get("Product_Name", "") or sku).strip()

    if packages_summary is not None and not packages_summary.empty:
        for _, row in packages_summary.iterrows():
            sku = str(row.get("Package_SKU", "") or "").strip()
            if sku:
                names[sku] = str(row.get("Package_Name", "") or sku).strip()

    return names


def build_cost_catalog(
    materials: dict,
    product_recipes: dict,
    products_summary: pd.DataFrame,
    package_compositions: dict,
    packages_summary: pd.DataFrame,
) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    names_by_sku = _build_cost_name_lookup(materials, products_summary, packages_summary)

    for _, row in products_summary.iterrows():
        sku = str(row.get("Product_SKU", "")).strip()
        name = row.get("Product_Name", sku)
        rows.append(
            {
                "System_SKU": sku,
                "System_Name": name,
                "Item_Type": "منتج",
                "COGS": resolve_component_cost(sku, materials, product_recipes, package_compositions),
                "COGS_Details": build_cost_details(
                    sku,
                    materials,
                    product_recipes,
                    package_compositions,
                    names_by_sku=names_by_sku,
                ),
            }
        )

    for _, row in packages_summary.iterrows():
        sku = str(row.get("Package_SKU", "")).strip()
        name = row.get("Package_Name", sku)
        rows.append(
            {
                "System_SKU": sku,
                "System_Name": name,
                "Item_Type": "بكج",
                "COGS": resolve_component_cost(sku, materials, product_recipes, package_compositions),
                "COGS_Details": build_cost_details(
                    sku,
                    materials,
                    product_recipes,
                    package_compositions,
                    names_by_sku=names_by_sku,
                ),
            }
        )

    catalog = pd.DataFrame(rows)
    if catalog.empty:
        return catalog

    catalog["Normalized_SKU"] = catalog["System_SKU"].map(normalize_sku)
    return catalog


def match_salla_prices_to_catalog(salla_df: pd.DataFrame, catalog_df: pd.DataFrame) -> pd.DataFrame:
    if catalog_df.empty:
        raise ValueError("No Zoho cost catalog is available. Upload Zoho files first.")

    salla_name_col = find_column(salla_df, SALLA_NAME_CANDIDATES, required=False)
    salla_sku_col = SALLA_MATCH_SKU_COL if SALLA_MATCH_SKU_COL in salla_df.columns else find_column(salla_df, SALLA_SKU_CANDIDATES, required=True)
    salla_id_col = find_column(salla_df, SALLA_ID_CANDIDATES, required=False)

    sku_index: dict[str, pd.Series] = {}
    for _, catalog_row in catalog_df.iterrows():
        sku = catalog_row.get("Normalized_SKU", "")
        if sku and sku not in sku_index:
            sku_index[sku] = catalog_row

    matched_rows = []

    for _, salla_row in salla_df.iterrows():
        salla_sku = normalize_sku(salla_row.get(salla_sku_col, ""))
        salla_name = salla_row.get(salla_name_col, "") if salla_name_col else ""
        match_row = None
        match_method = "SKU"
        match_note = ""

        if salla_sku and salla_sku in sku_index:
            match_row = sku_index[salla_sku]
        elif not salla_sku:
            match_note = "SKU سلة مفقود أو صفر"
        else:
            match_note = "SKU سلة غير موجود في بيانات Zoho"

        output = salla_row.to_dict()
        output["Salla_Product_ID"] = salla_row.get(salla_id_col, "") if salla_id_col else ""
        output["Salla_Name"] = salla_name
        output["Salla_SKU"] = salla_sku
        output["System_SKU"] = match_row.get("System_SKU", "") if match_row is not None else ""
        output["System_Name"] = match_row.get("System_Name", "") if match_row is not None else ""
        output["Item_Type"] = match_row.get("Item_Type", "") if match_row is not None else ""
        output["COGS"] = match_row.get("COGS", pd.NA) if match_row is not None else pd.NA
        output["COGS_Details"] = match_row.get("COGS_Details", "") if match_row is not None else ""
        output["Match_Status"] = "Matched" if match_row is not None else "Not Matched"
        output["Match_Method"] = match_method
        output["Match_Note"] = match_note
        matched_rows.append(output)

    return pd.DataFrame(matched_rows)


def _channel_to_dict(channel: Any) -> dict[str, float]:
    if hasattr(channel, "__dataclass_fields__"):
        data = asdict(channel)
    elif isinstance(channel, dict):
        data = channel
    else:
        data = getattr(channel, "__dict__", {})

    return {
        "platform_pct": float(data.get("platform_pct", 0) or 0),
        "payment_pct": float(data.get("payment_pct", 0) or 0),
        "marketing_pct": float(data.get("marketing_pct", 0) or 0),
        "opex_pct": float(data.get("opex_pct", 0) or 0),
        "vat_rate": float(data.get("vat_rate", 0.15) or 0.15),
    }


def calculate_salla_price_review(
    salla_df: pd.DataFrame,
    catalog_df: pd.DataFrame,
    channel: Any,
    min_margin_pct: float = 15.0,
) -> tuple[pd.DataFrame, dict[str, int]]:
    prepared_df, _ = add_customer_final_price_columns(salla_df, vat_rate=0.15)
    matched_df = match_salla_prices_to_catalog(prepared_df, catalog_df)

    final_price = pd.to_numeric(matched_df[FINAL_INCL_VAT_COL], errors="coerce")
    cogs = pd.to_numeric(matched_df["COGS"], errors="coerce")

    channel_dict = _channel_to_dict(channel)
    vat_rate = channel_dict["vat_rate"]
    shipping = float(getattr(channel, "shipping_fixed", 0) or 0)
    preparation = float(getattr(channel, "preparation_fee", 0) or 0)
    free_shipping_threshold = float(getattr(channel, "free_shipping_threshold", 0) or 0)
    custom_fees = getattr(channel, "custom_fees", {}) or {}

    review_rows = []
    for idx, row in matched_df.iterrows():
        row_dict = row.to_dict()
        row_final_price = final_price.iloc[idx]
        row_cogs = cogs.iloc[idx]

        row_dict.update(
            {
                "Active_On_Store": bool(pd.notna(row_final_price) and row_final_price > 0),
                "Net_Price_Excl_VAT": pd.NA,
                "Sales_VAT": pd.NA,
                "Shipping_Fee": pd.NA,
                "Preparation_Fee": pd.NA,
                "Platform_Fee": pd.NA,
                "Payment_Fee": pd.NA,
                "Marketing_Fee": pd.NA,
                "Opex_Fee": pd.NA,
                "Custom_Fees": pd.NA,
                "Channel_Fees_Total": pd.NA,
                "Total_Costs_And_Fees": pd.NA,
                "Profit": pd.NA,
                "Margin_%": pd.NA,
                "Fees_Formula": "",
                "Profit_Formula": "",
                "Breakeven_Price_Incl_VAT": pd.NA,
                "Price_Gap_To_Breakeven": pd.NA,
                "Review_Status": "تحتاج مراجعة",
                "Review_Note": "",
            }
        )

        if row_dict["Match_Status"] != "Matched":
            row_dict["Review_Note"] = row_dict.get("Match_Note") or "SKU سلة غير موجود في بيانات Zoho"
            review_rows.append(row_dict)
            continue
        if pd.isna(row_final_price) or row_final_price <= 0:
            row_dict["Review_Note"] = "سعر سلة النهائي مفقود"
            review_rows.append(row_dict)
            continue
        if pd.isna(row_cogs) or row_cogs <= 0:
            row_dict["Review_Note"] = "التكلفة مفقودة أو صفر"
            review_rows.append(row_dict)
            continue

        try:
            breakdown = calculate_price_breakdown(
                cogs=float(row_cogs),
                channel_fees=channel_dict,
                shipping=shipping,
                preparation=preparation,
                discount_rate=0.0,
                vat_rate=vat_rate,
                free_shipping_threshold=free_shipping_threshold,
                custom_fees=custom_fees,
                price_with_vat=float(row_final_price),
            )
            margin_pct = float(breakdown["margin_pct"] * 100)
            profit = float(breakdown["profit"])
            breakeven_price = float(breakdown["breakeven_price"])
            net_price = float(breakdown["net_price"])
            channel_fees_total = float(breakdown["total_costs_fees"] - breakdown["cogs"])
            sales_vat = float(row_final_price) - net_price
            shipping_fee = float(breakdown["shipping_fee"])
            preparation_fee = float(breakdown["preparation_fee"])
            platform_fee = float(breakdown["platform_fee"])
            payment_fee = float(breakdown.get("payment_fee", 0))
            marketing_fee = float(breakdown["marketing_fee"])
            opex_fee = float(breakdown["admin_fee"])
            custom_fees_total = float(breakdown.get("custom_fees_total", 0))

            if profit < 0:
                review_status = "خسارة"
            elif margin_pct < min_margin_pct:
                review_status = "هامش منخفض"
            else:
                review_status = "مناسب"

            row_dict.update(
                {
                    "Net_Price_Excl_VAT": net_price,
                    "Sales_VAT": sales_vat,
                    "Shipping_Fee": shipping_fee,
                    "Preparation_Fee": preparation_fee,
                    "Platform_Fee": platform_fee,
                    "Payment_Fee": payment_fee,
                    "Marketing_Fee": marketing_fee,
                    "Opex_Fee": opex_fee,
                    "Custom_Fees": custom_fees_total,
                    "Channel_Fees_Total": channel_fees_total,
                    "Total_Costs_And_Fees": breakdown["total_costs_fees"],
                    "Profit": profit,
                    "Margin_%": margin_pct,
                    "Fees_Formula": (
                        f"{shipping_fee:.2f} + {preparation_fee:.2f} + {platform_fee:.2f} + "
                        f"{payment_fee:.2f} + {marketing_fee:.2f} + {opex_fee:.2f} + "
                        f"{custom_fees_total:.2f} = {channel_fees_total:.2f}"
                    ),
                    "Profit_Formula": (
                        f"{float(row_final_price):.2f} - {sales_vat:.2f} - "
                        f"{float(row_cogs):.2f} - {channel_fees_total:.2f} = {profit:.2f}"
                    ),
                    "Breakeven_Price_Incl_VAT": breakeven_price,
                    "Price_Gap_To_Breakeven": float(row_final_price) - breakeven_price,
                    "Review_Status": review_status,
                    "Review_Note": "",
                }
            )
        except Exception as exc:
            row_dict["Review_Note"] = f"تعذر حساب الربحية: {exc}"

        review_rows.append(row_dict)

    review_df = pd.DataFrame(review_rows)
    summary = {
        "total_rows": int(len(review_df)),
        "active_rows": int(review_df["Active_On_Store"].sum()) if "Active_On_Store" in review_df.columns else 0,
        "matched_rows": int((review_df["Match_Status"] == "Matched").sum()),
        "unmatched_rows": int((review_df["Match_Status"] != "Matched").sum()),
        "profitable_rows": int((review_df["Review_Status"] == "مناسب").sum()),
        "low_margin_rows": int((review_df["Review_Status"] == "هامش منخفض").sum()),
        "loss_rows": int((review_df["Review_Status"] == "خسارة").sum()),
        "needs_review_rows": int((review_df["Review_Status"] == "تحتاج مراجعة").sum()),
    }
    return review_df, summary
